import streamlit as st
import pandas as pd
import re
import io
from fpdf import FPDF
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# --- Bloco 1: L√≥gica Principal da Concilia√ß√£o ---

def gerar_chave_padronizada(texto_conta):
    """
    Padroniza a cria√ß√£o da chave para DE-PARA e Extratos.
    1. Extrai apenas os d√≠gitos.
    2. Pega os √∫ltimos 7 d√≠gitos.
    3. Garante que a chave tenha SEMPRE 7 d√≠gitos.
    """
    if isinstance(texto_conta, str):
        parte_numerica = re.sub(r'\D', '', texto_conta)
        ultimos_7_digitos = parte_numerica[-7:]
        return ultimos_7_digitos.zfill(7)
    return None

def gerar_chave_contabil(texto_conta):
    """
    Extrai a chave do campo 'Domic√≠lio banc√°rio' e a padroniza para 7 d√≠gitos.
    """
    if not isinstance(texto_conta, str):
        return None
    try:
        partes = texto_conta.split('-')
        if len(partes) > 2:
            parte_conta = partes[2]
            conta_numerica = re.sub(r'\D', '', parte_conta)
            ultimos_7_digitos = conta_numerica[-7:]
            return ultimos_7_digitos.zfill(7)
    except (IndexError, AttributeError):
        return None
    return None

def carregar_depara():
    """Carrega o arquivo DE-PARA e padroniza as chaves."""
    try:
        df_depara = pd.read_excel(
            "depara/DEPARA_CONTAS BANC√ÅRIAS_CEF.xlsx",
            sheet_name="2025_JUNHO (2)",
            dtype=str
        )
        df_depara.columns = ['Conta Antiga', 'Conta Nova']
        df_depara['Chave Antiga'] = df_depara['Conta Antiga'].apply(gerar_chave_padronizada)
        df_depara['Chave Nova'] = df_depara['Conta Nova'].apply(gerar_chave_padronizada)
        return df_depara
    except FileNotFoundError:
        st.warning("Aviso: Arquivo DE-PARA 'depara/DEPARA_CONTAS BANC√ÅRIAS_CEF.xlsx' n√£o encontrado. A tradu√ß√£o de contas n√£o ser√° aplicada.")
        return pd.DataFrame()

def processar_relatorio_contabil(arquivo_carregado, df_depara):
    """L√™ o relat√≥rio cont√°bil e aplica a tradu√ß√£o DE-PARA."""
    df = pd.read_csv(arquivo_carregado, encoding='latin-1', sep=';', header=1)

    df['Chave Primaria'] = df['Domic√≠lio banc√°rio'].apply(gerar_chave_contabil)
    
    df.dropna(subset=['Chave Primaria'], inplace=True)
    df = df[df['Chave Primaria'] != '']

    if not df_depara.empty:
        df_depara_map = df_depara.copy()
        df_depara_map['Chave Antiga'] = df_depara_map['Chave Antiga'].astype(str)
        df['Chave Primaria'] = df['Chave Primaria'].astype(str)
        
        mapa_depara = df_depara_map.set_index('Chave Antiga')['Chave Nova'].to_dict()
        df['Chave Primaria'] = df['Chave Primaria'].replace(mapa_depara)

    df['Saldo Final'] = pd.to_numeric(
        df['Saldo Final'].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False),
        errors='coerce'
    ).fillna(0)

    df_pivot = df.pivot_table(index='Chave Primaria', columns='Conta cont√°bil', values='Saldo Final', aggfunc='sum').reset_index()

    rename_dict = {c: 'Saldo_Corrente_Contabil' for c in df_pivot.columns if '111111901' in str(c)}
    rename_dict.update({c: 'Saldo_Aplicado_Contabil' for c in df_pivot.columns if '111115001' in str(c)})
    df_pivot.rename(columns=rename_dict, inplace=True)

    mapa_conta = df[['Chave Primaria', 'Domic√≠lio banc√°rio']].drop_duplicates().set_index('Chave Primaria')
    df_final = df_pivot.join(mapa_conta, on='Chave Primaria')

    if 'Saldo_Corrente_Contabil' not in df_final.columns:
        df_final['Saldo_Corrente_Contabil'] = 0
    if 'Saldo_Aplicado_Contabil' not in df_final.columns:
        df_final['Saldo_Aplicado_Contabil'] = 0

    return df, df_final

def processar_extrato_bb_bruto_csv(caminho_arquivo):
    """
    L√™ e transforma o arquivo .csv bruto do Banco do Brasil.
    Esta vers√£o foi corrigida para lidar com valores num√©ricos onde os dois 
    √∫ltimos d√≠gitos representam os centavos (ex: '12345' se torna 123.45).
    """
    df = pd.read_csv(caminho_arquivo, sep=',', encoding='latin-1', dtype=str)
    df.rename(columns={
        'Saldo em conta': 'Saldo_Corrente_Extrato',
        'Saldo investido': 'Saldo_Aplicado_Extrato'
    }, inplace=True)
    df['Chave Primaria'] = df['Conta'].apply(gerar_chave_padronizada)
    
      # --- IN√çCIO DO BLOCO MODIFICADO ---
    # Converte as colunas de saldo, tratando os valores como inteiros
    # e dividindo por 1000 para obter os centavos.
    for col in ['Saldo_Corrente_Extrato', 'Saldo_Aplicado_Extrato', 'Saldo total']:
        if col in df.columns:
            # 1. Converte a coluna para texto (garantia).
            # 2. Remove quaisquer caracteres n√£o num√©ricos (como R$, pontos ou v√≠rgulas).
            # 3. Converte o texto limpo para um n√∫mero.
            # 4. Divide por 1000 para ajustar os centavos.
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace(r'\D', '', regex=True), # Remove tudo que n√£o for d√≠gito
                errors='coerce'
            ).fillna(0) / 1000
    # --- FIM DO BLOCO MODIFICADO ---
            
    # Garante que as colunas existam, caso n√£o venham no arquivo original
    if 'Saldo_Corrente_Extrato' not in df.columns:
        df['Saldo_Corrente_Extrato'] = 0
    if 'Saldo_Aplicado_Extrato' not in df.columns:
        df['Saldo_Aplicado_Extrato'] = 0
            
    return df

def processar_extrato_cef_bruto(caminho_arquivo):
    """L√™ o arquivo .cef da Caixa."""
    with open(caminho_arquivo, 'r', encoding='latin-1') as f:
        cef_content = f.readlines()
    header_line_index = -1
    for i, line in enumerate(cef_content):
        if line.strip().startswith("Conta Vinculada;"):
            header_line_index = i
            break
    if header_line_index == -1: return pd.DataFrame()
    data_io = io.StringIO("".join(cef_content[header_line_index:]))
    df = pd.read_csv(data_io, sep=';', dtype=str)
    df['Chave Primaria'] = df['Conta Vinculada'].apply(gerar_chave_padronizada)
    df.rename(columns={
        'Saldo Conta Corrente (R$)': 'Saldo_Corrente_Extrato',
        'Saldo Aplicado (R$)': 'Saldo_Aplicado_Extrato'
    }, inplace=True)

    # --- IN√çCIO DA SE√á√ÉO DE TRATAMENTO NUM√âRICO ---
    # Esta l√≥gica j√° est√° correta e robusta.
    for col in ['Saldo_Corrente_Extrato', 'Saldo_Aplicado_Extrato']:
        if col in df.columns:
            # A linha abaixo j√° faz a limpeza de '.' e a substitui√ß√£o de ',' por '.'
            df[col] = pd.to_numeric(df[col].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False), errors='coerce').fillna(0)
    # --- FIM DA SE√á√ÉO DE TRATAMENTO NUM√âRICO ---
            
    if 'Saldo_Corrente_Extrato' not in df.columns:
        df['Saldo_Corrente_Extrato'] = 0
    if 'Saldo_Aplicado_Extrato' not in df.columns:
        df['Saldo_Aplicado_Extrato'] = 0
    return df

def realizar_conciliacao(df_contabil, df_extrato_unificado):
    df_contabil_pivot = df_contabil[['Chave Primaria', 'Domic√≠lio banc√°rio', 'Saldo_Corrente_Contabil', 'Saldo_Aplicado_Contabil']]
    df_extrato_pivot = df_extrato_unificado.groupby('Chave Primaria').agg({
        'Saldo_Corrente_Extrato': 'sum',
        'Saldo_Aplicado_Extrato': 'sum'
    }).reset_index()
    
    df_contabil_pivot['Chave Primaria'] = df_contabil_pivot['Chave Primaria'].astype(str)
    df_extrato_pivot['Chave Primaria'] = df_extrato_pivot['Chave Primaria'].astype(str)

    df_final = pd.merge(df_contabil_pivot, df_extrato_pivot, on='Chave Primaria', how='inner')
    if df_final.empty: return pd.DataFrame()
    df_final.rename(columns={'Domic√≠lio banc√°rio': 'Conta Banc√°ria'}, inplace=True)
    df_final['Diferenca_Movimento'] = df_final['Saldo_Corrente_Contabil'] - df_final['Saldo_Corrente_Extrato']
    df_final['Diferenca_Aplicacao'] = df_final['Saldo_Aplicado_Contabil'] - df_final['Saldo_Aplicado_Extrato']
    df_final = df_final.set_index('Conta Banc√°ria')
    df_final = df_final[['Saldo_Corrente_Contabil', 'Saldo_Corrente_Extrato', 'Diferenca_Movimento','Saldo_Aplicado_Contabil', 'Saldo_Aplicado_Extrato', 'Diferenca_Aplicacao']]
    df_final.columns = pd.MultiIndex.from_tuples([
        ('Conta Movimento', 'Saldo Cont√°bil'), ('Conta Movimento', 'Saldo Extrato'), ('Conta Movimento', 'Diferen√ßa'),
        ('Aplica√ß√£o Financeira', 'Saldo Cont√°bil'), ('Aplica√ß√£o Financeira', 'Saldo Extrato'), ('Aplica√ß√£o Financeira', 'Diferen√ßa')
    ], names=['Grupo', 'Item'])
    return df_final

# --- Bloco 2: Fun√ß√µes para Gera√ß√£o de Arquivos ---
@st.cache_data
def to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=True, sheet_name='Conciliacao', startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['Conciliacao']
        font_header = Font(bold=True, color="FFFFFF")
        align_header = Alignment(horizontal='center', vertical='center')
        fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        number_format_br = '#,##0.00'
        worksheet.merge_cells('B1:D1'); cell_movimento = worksheet['B1']; cell_movimento.value = 'Conta Movimento'; cell_movimento.font = font_header; cell_movimento.alignment = align_header; cell_movimento.fill = fill_header
        worksheet.merge_cells('E1:G1'); cell_aplicacao = worksheet['E1']; cell_aplicacao.value = 'Aplica√ß√£o Financeira'; cell_aplicacao.font = font_header; cell_aplicacao.alignment = align_header; cell_aplicacao.fill = fill_header
        for row in worksheet['A2:G2']:
            for cell in row: cell.font = Font(bold=True); cell.alignment = Alignment(horizontal='center', vertical='center')
        for col_idx, col in enumerate(worksheet.columns, 1):
            max_length = 0; column_letter = get_column_letter(col_idx)
            for cell_idx, cell in enumerate(col, 0):
                if cell_idx > 0: cell.border = border_thin
                if cell_idx > 1:
                    if col_idx == 1: cell.alignment = Alignment(horizontal='left', vertical='center')
                    else: cell.number_format = number_format_br; cell.alignment = Alignment(horizontal='right', vertical='center')
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2); worksheet.column_dimensions[column_letter].width = adjusted_width
    return output.getvalue()

class PDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 12); self.cell(0, 8, 'Prefeitura da Cidade do Rio de Janeiro', 0, 1, 'C'); self.set_font('Arial', '', 11); self.cell(0, 8, 'Controladoria Geral do Munic√≠pio', 0, 1, 'C'); self.set_font('Arial', 'B', 10); self.cell(0, 8, 'Relat√≥rio de Concilia√ß√£o de Saldos Banc√°rios', 0, 1, 'C'); self.ln(5)
    def footer(self):
        self.set_y(-15); self.set_font('Arial', 'I', 8); self.cell(0, 10, f'P√°gina {self.page_no()}', 0, 0, 'C')
    def create_table(self, data):
        self.set_font('Arial', '', 7); line_height = self.font_size * 2.5; col_width = 30
        self.set_font('Arial', 'B', 8); index_name = data.index.name if data.index.name else 'ID'; self.cell(40, line_height, index_name, 1, 0, 'C'); self.cell(col_width * 3, line_height, 'Conta Movimento', 1, 0, 'C'); self.cell(col_width * 3, line_height, 'Aplica√ß√£o Financeira', 1, 0, 'C'); self.ln(line_height)
        self.set_font('Arial', 'B', 7); self.cell(40, line_height, '', 1, 0, 'C'); sub_headers = ['Saldo Cont√°bil', 'Saldo Extrato', 'Diferen√ßa'];
        for _ in range(2):
            for sub_header in sub_headers: self.cell(col_width, line_height, sub_header, 1, 0, 'C')
        self.ln(line_height)
        self.set_font('Arial', '', 6); formatted_data = data.copy()
        for col_tuple in formatted_data.columns: formatted_data[col_tuple] = formatted_data[col_tuple].apply(lambda x: f'{x:,.2f}'.replace(",", "X").replace(".", ",").replace("X", "."))
        for index, row in formatted_data.iterrows():
            display_index = str(index); self.cell(40, line_height, display_index, 1, 0, 'L')
            for item in row: self.cell(col_width, line_height, str(item), 1, 0, 'R')
            self.ln(line_height)

def create_pdf(df):
    pdf = PDF('L', 'mm', 'A4'); pdf.add_page(); pdf.create_table(df); return bytes(pdf.output())

# --- Bloco 3: Interface Web com Streamlit ---
st.set_page_config(page_title="Concilia√ß√£o Banc√°ria", layout="wide", page_icon="üè¶")
st.title("üè¶ Prefeitura da Cidade do Rio de Janeiro"); st.header("Controladoria Geral do Munic√≠pio"); st.markdown("---"); st.subheader("Concilia√ß√£o de Saldos Banc√°rios e Cont√°beis")

meses = {1: "janeiro", 2: "fevereiro", 3: "mar√ßo", 4: "abril", 5: "maio", 6: "junho", 7: "julho", 8: "agosto", 9: "setembro", 10: "outubro", 11: "novembro", 12: "dezembro"}
ano_atual = datetime.now().year
opcoes_meses_formatadas = [f"{nome.capitalize()} {ano}" for ano in range(ano_atual, ano_atual + 2) for mes, nome in meses.items()]
try:
    index_padrao = opcoes_meses_formatadas.index(f"{meses[datetime.now().month].capitalize()} {ano_atual}")
except ValueError:
    index_padrao = 0
st.selectbox("Selecione o M√™s da Concilia√ß√£o:", options=opcoes_meses_formatadas, index=index_padrao, key='mes_selecionado')

st.sidebar.header("Carregar Relat√≥rio Cont√°bil")
contabilidade_bruto = st.sidebar.file_uploader(f"Selecione o seu Relat√≥rio Cont√°bil Bruto de {st.session_state.mes_selecionado}", type=['csv'])

if st.sidebar.button("Conciliar Agora"):
    if contabilidade_bruto is not None:
        with st.spinner("Processando..."):
            try:
                partes_mes = st.session_state.mes_selecionado.lower().split()
                mes_ano = f"{partes_mes[0]}_{partes_mes[1]}"
                
                df_depara = carregar_depara()
                st.session_state['audit_depara'] = df_depara
                
                extratos_encontrados = []
                st.session_state['audit_bb'] = None
                st.session_state['audit_cef'] = None
                
                try:
                    caminho_bb = f"extratos_consolidados/extrato_bb_{mes_ano}.csv"
                    df_bb = processar_extrato_bb_bruto_csv(caminho_bb)
                    extratos_encontrados.append(df_bb)
                    st.session_state['audit_bb'] = df_bb
                except FileNotFoundError:
                    st.warning(f"Aviso: Extrato do BB (.csv) para {st.session_state.mes_selecionado} n√£o encontrado.")
                
                try:
                    caminho_cef = f"extratos_consolidados/extrato_cef_{mes_ano}.cef"
                    df_cef = processar_extrato_cef_bruto(caminho_cef)
                    extratos_encontrados.append(df_cef)
                    st.session_state['audit_cef'] = df_cef
                except FileNotFoundError:
                    st.warning(f"Aviso: Extrato da CEF (.cef) para {st.session_state.mes_selecionado} n√£o encontrado.")

                extratos_encontrados = [df for df in extratos_encontrados if df is not None and not df.empty]

                if not extratos_encontrados:
                    st.error("Nenhum arquivo de extrato v√°lido foi encontrado no reposit√≥rio para o m√™s selecionado.")
                    st.session_state['df_resultado'] = None
                else:
                    df_extrato_unificado = pd.concat(extratos_encontrados, ignore_index=True)
                    df_contabil_raw_audit, df_contabil_processado = processar_relatorio_contabil(contabilidade_bruto, df_depara)
                    st.session_state['audit_contabil'] = df_contabil_raw_audit
                    df_resultado_final = realizar_conciliacao(df_contabil_processado, df_extrato_unificado)
                    st.success("Concilia√ß√£o Conclu√≠da com Sucesso!")
                    st.session_state['df_resultado'] = df_resultado_final
            except Exception as e:
                st.error(f"Ocorreu um erro durante o processamento: {e}")
                st.session_state['df_resultado'] = None
    else:
        st.sidebar.warning("Por favor, carregue o seu arquivo de relat√≥rio cont√°bil.")

if 'df_resultado' in st.session_state and st.session_state['df_resultado'] is not None:
    resultado = st.session_state['df_resultado']
    if isinstance(resultado, pd.DataFrame):
        if resultado.empty:
            st.warning("Processamento conclu√≠do. Nenhuma conta correspondente foi encontrada entre o relat√≥rio cont√°bil e os extratos para gerar um relat√≥rio de concilia√ß√£o.")
        else:
            st.header("Resultado da Concilia√ß√£o Consolidada")
            df_para_mostrar = resultado[(resultado[('Conta Movimento', 'Diferen√ßa')].abs() > 0.01) | (resultado[('Aplica√ß√£o Financeira', 'Diferen√ßa')].abs() > 0.01)].copy()
            if df_para_mostrar.empty:
                st.success("‚úÖ √ìtima not√≠cia! Nenhuma diverg√™ncia encontrada.")
            else:
                st.write("A tabela abaixo mostra apenas as contas com diverg√™ncia de saldo.")
                formatters = {col: (lambda x: f'{x:,.2f}'.replace(",", "X").replace(".", ",").replace("X", ".")) for col in resultado.columns}
                st.dataframe(df_para_mostrar.style.format(formatter=formatters))
            st.header("Download do Relat√≥rio Completo")
            col1, col2, col3 = st.columns(3)
            with col1:
                df_csv = resultado.copy(); df_csv.columns = [' - '.join(map(str,col)).strip() for col in df_csv.columns.values]; st.download_button("Baixar em CSV", df_csv.to_csv(index=True, sep=';', decimal=',').encode('utf-8-sig'), 'relatorio_consolidado.csv', 'text/csv')
            with col2:
                st.download_button("Baixar em Excel", to_excel(resultado), 'relatorio_consolidado.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            with col3:
                st.download_button("Baixar em PDF", create_pdf(resultado), 'relatorio_consolidado.pdf', 'application/pdf')
        st.markdown("---")
        with st.expander("Clique aqui para auditar os dados de origem"):
            
            st.subheader("Auditoria do Arquivo DE-PARA")
            if 'audit_depara' in st.session_state and st.session_state['audit_depara'] is not None:
                df_audit_view = st.session_state['audit_depara'].copy()
                if not df_audit_view.empty:
                    df_audit_view = df_audit_view[['Conta Antiga', 'Chave Antiga', 'Conta Nova', 'Chave Nova']]
                    df_audit_view.columns = [
                        'Conta Original (Antiga)', 
                        'Chave Gerada (Antiga)', 
                        'Conta Original (Nova)', 
                        'Chave Gerada (Nova)'
                    ]
                    st.dataframe(df_audit_view)
            
            st.subheader("Auditoria do Relat√≥rio Cont√°bil (com Chave Prim√°ria)")
            if 'audit_contabil' in st.session_state and st.session_state['audit_contabil'] is not None:
                st.dataframe(st.session_state['audit_contabil'])

            st.subheader("Auditoria do Extrato do Banco do Brasil (com Chave Prim√°ria)")
            if 'audit_bb' in st.session_state and st.session_state['audit_bb'] is not None:
                st.dataframe(st.session_state['audit_bb'])

            st.subheader("Auditoria do Extrato da Caixa Econ√¥mica (com Chave Prim√°ria)")
            if 'audit_cef' in st.session_state and st.session_state['audit_cef'] is not None:
                st.dataframe(st.session_state['audit_cef'])
